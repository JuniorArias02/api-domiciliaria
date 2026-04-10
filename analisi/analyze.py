import openpyxl
import json

def analyze_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        report = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(max_row=5, values_only=True))
            if not rows:
                report[sheet_name] = {"headers": [], "rows": []}
                continue
            
            headers = rows[0]
            report[sheet_name] = {
                "headers": headers,
                "rows": rows[1:]
            }
        print(json.dumps(report, indent=4))
    except Exception as e:
        print(json.dumps({"error": str(e)}))

if __name__ == "__main__":
    analyze_xlsx("consulta enero a marzo.xlsx")
